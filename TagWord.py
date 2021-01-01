# -*- coding: utf-8 -*-

# tag_word : after extract word data from paper, taging all of the data
# requirement :
# 1. 포맷에 맞춘 엑셀파일(2개의 row로 쪼개야함 )
# 2. 단락 넘어갈 때 '-' 로 된 단어 미리 수정해놔야함
# 3. 데이터 중복화 적용 해줘야 함
# date : 2020/12/30

import openpyxl
from os import listdir
from os.path import isfile, join

# Data tagging rule
mp = ['property', 'properties', 'conduction', 'driving', 'frequency', 'frequencies', 'absorbance', 'space_charge_limited', 'scattering', 'phase', 'valence', 'electronic', 'doping', 'force', 'vibrational', 'opacity', 'mass', 'holes', 'solid-state', 'thermochemical', 'doped', 'endothermic', 'energy', 'electric', 'excited', 'barrier', 'physical', 'electrochemical', 'intrinsic', 'transmittance', 'field', 'chemical', 'semiconductor', 'HOMO', 'metallic', 'charge', 'electrons', 'bands', 'electron', 'atomic', 'transport', 'LUMO', 'insulator', 'electrically', 'trapping', 'detrapping', 'stretching', 'electromagnetic', 'metal', 'carbon-rich', 'chemically']
sp_dev = ['/']
ds = ['structure', 'TE', 'BE', 'electrode', 'conventional', 'electrolyte', 'dielectric', 'buffer', 'active', 'top', 'bottom', 'layer', 'substrate', 'thickness', 'thick', 'tox', 'multistack', 'vertical', 'channel', 'Cluster', 'cluster', 'MIM', 'cross-point', 'nanowire', 'nanodot', 'nanomesh', 'X-point', 'oxram', 'CBRAM', 'terminal', 'surface', 'PMC', 'ReRAM', 'width', 'pitch', 'film']
da = ['application', 'selector', 'memristors', 'memristive', 'Memory', 'neuromorphic', 'switch', 'mulit-level', 'MLC', 'memories', 'Neural', 'NVM', 'storage', 'floating_gate', 'nonvolatile', 'transistors']

p = ["performance"]
pp = ["power", "consumption"]
pc = ['current', 'Ion', 'Iset', 'Ioff', 'leakage', 'Ileak', 'Ireset', 'compliance', 'Iprog', 'Ierase', 'Icc', 'IHRS', 'ILRS']
pv = ['voltage', 'Vset', 'Vreset', 'Vforming', 'Vread', 'Verase', 'Vprogram']
pr = ['resistance', 'resistive', 'conductivity', 'high_resistance', 'low-resistance', 'HRS', 'LRS', 'RH', 'RL', 'Roff', 'Ron', 'resistance-state']
po = ['operating', 'Forming-free', 'semi-forming', 'unipolar', 'bipolar', 'complementary', 'ON', 'OFF', 'operation', 'Program', 'erase', 'read', 'write', 'rupture', 'fast', 'slow', 'forming', 'current–voltage', 'I–V','formation', 'Electroforming', 'switching', 'breakdown', 'high', 'low', 'set', 'reset', 'positive', 'negative', 'threshold', 'sweep', 'dissolution', 'polarity']
ps = ['speed']
pe = ['endurance', 'cycle', 'cycles', 'cycling', 'cyclability']
pt = ['retention', 'lifetime']
pab = ['reliability', 'stability', 'variability', 'disturbance', 'uniformity', 'dispersion', 'distributions', 'cumulative', 'Fluctuation', 'deviation', 'window', 'non-uniformity', 'uniform', 'reproducible', 'probabilities']
plec = ['selectivity', 'ratio', 'Non-linearity']

mc = ['mechanism', 'VCM', 'ECM', 'interface-switching', 'filament', 'thermal-chemical', 'mechanism', 'URS', 'BRS', 'CRS', 'frenkel', 'electrochemical', 'path', 'TCM', 'Schottky', 'ohmic', 'Poole', 'precipitation', 'SCLC', 'oxidized', 'reduction']
e = ['environment', 'humidity', 'temperature', 'pressure', 'time', 'heat', 'air', 'dry']
s = ['synthesis', 'RF-sputtering', 'sputtering', 'e-beam', 'evaporator', 'spin-coated', 'sputtering', 'annealing', 'laser', 'lithography', 'photolithography', 'etching', 'etched', 'patterning', 'ALD', 'CVD', 'PVD', 'PLD', 'plasma', 'deposition', 'process', 'solution', 'self-assembly', 'drying', 'thermal', 'angled', 'beam', 'vapor', 'printing']

u = ["kg", "nm", "mm", "mA", "V", "Ω", "lm","µm", "µ", "ppm", "L", "C"]

chemi = ['H', 'He', 'NiO', 'Li', 'N', 'O', 'F', 'Ne', 'Na', 'Mg', 'Al', 'Si', 'P', 'S', 'Cl', 'Ar', 'Ca', 'Sc', 'Ti', 'Cr', 'Fe', 'Co', 'Ni', 'Mn', 'Fe', 'Co', 'Ni', 'Cu', 'Zn', 'Ga', 'Ge', 'Se', 'Br', 'Kr', 'Rb', 'Sr', 'Y', 'Zr', 'Nb', 'Mo', 'Tc', 'Ru', 'Rh', 'Pd', 'Ag', 'Cd', 'Sn', 'Sb', 'Te', 'Xe', 'Cs', 'Ba', 'Hf', 'Ta', 'Re', 'Os', 'Ir', 'Pt', 'Au', 'Hg', 'Tl', 'Pb', 'Bi', 'Po', 'Rn', 'Fr', 'Ra', 'Rf', 'Db', 'Sg', 'Bh', 'Hs', 'Mt', 'Ds', 'Rg', 'Cn']
# caution = ['As', 'In', 'V', 'B', 'C', 'I', 'K']
# 온도 단위 = chemi -> ['C','K']

# dictionary 형태로 다시 묶음
data_set = {'mp': mp, 'sp_dev': sp_dev, 'ds': ds, 'da': da,
            'p' : p, 'pp': pp, 'pc': pc,
            'pv': pv, 'pr': pr, 'po': po, 'ps': ps, 'pe': pe, 'pt': pt,
            'pab': pab, 'plec': plec, 'mc': plec, 'e': e, 's': s, 'u': u, 'chemi': chemi}

def check_number(number):
    try:
        float(number)
        return True
    except: return False

def word_check(word,data_set):
    checking = 'o'
    word = str(word).replace('^', '').replace(',', '')
    for tag, checked_list in data_set.items():
        upper_list = [x.upper() for x in checked_list]
        if check_number(word):
            checking = 'n'
            break
        elif word.upper() in upper_list:
            checking = tag
            break
        # 숫자 체크 (음수, 소수 가능)
        # 자연상수는?
    return checking

# 절대 경로 / 상대 경로
AB_DIR = '/Users/SBJ/Desktop/kriss/Data/Text/Word'
DIR = 'Data/Text/Word/'

# Data/Text/Word 에 있는 파일을 리스트로 생성
file_list = [f for f in listdir(AB_DIR) if isfile(join(AB_DIR, f))]
# print(file_list)

wb = openpyxl.load_workbook(DIR + file_list[0])
sheet = wb.active

# sheet의 row 개수 만큼 검사
max_row = sheet.max_row + 1

for i in range(1, max_row):
    sheet.cell(i, 2).value = word_check(sheet.cell(i, 1).value, data_set)

wb.save(DIR + file_list[0])