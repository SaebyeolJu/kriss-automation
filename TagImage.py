# -*- coding: utf-8 -*-
import os
import openpyxl
import natsort
import sys
from os import listdir
from os.path import isfile, join
from os import rename, listdir

# 두께 비례식
def thickness():
    stand = float(input('기준 단위(nm) : '))
    stand_cm = float(input('기준 단위 측정 cm : '))
    mat_cm = float(input('구해야 할 소재 cm : '))
    mat = (stand * mat_cm) / stand_cm
    return str(mat)

def curve(paper_name, image_number, DIR, img_file_name, figure_number):
    print('IV curve 그래프')
    wb = openpyxl.load_workbook(DIR + '/image.xlsx')
    s_curve = wb['s_curve']

    info = ['file_name', 'figure_name', 'x-axis', 'y-axis', 'device',
            'set voltage (V)', 'off voltage (V)',
            'on_current (A)', 'off_current (A)', 'unipolar/bipolar']

    for i in range(0, len(info)):
        s_curve.cell(i+1, 1).value = info[i]

    curve_info = ['NAN'] * 10
    curve_info[1] = f'figure.' + input('Figure 이름 : ')

    fig_check = input('이전 figure와 다른 그래프 / 같은 figure인가요?(y/n) : ')
    if fig_check == 'y' : figure_number += 1
    else : figure_number = 1

    i_number = image_number
    image_name = f'{paper_name}.IV-curve{i_number}_{figure_number}'
    curve_info[0] = image_name

    x_first = input('x-axis first : ')
    x_last = input('x-axis last : ')
    curve_info[2] = f'[{x_first},{x_last}]'

    y_first = input('y-axis first : ')
    y_last = input('y-axis last : ')
    curve_info[3] = f'[{y_first},{y_last}]'

    # device
    curve_info[4] = input('device : ')

    # voltage
    curve_info[5] = input('set voltage : ')
    curve_info[6] = input('reset voltage : ')

    # current
    curve_info[7] = input('on current : ')
    curve_info[8] = input('off current : ')

    curve_info[9] = input('unipolar / bipolar?(1 / 2) : ')
    if curve_info[9] == '1': curve_info[9] = 'unipolar'
    else: curve_info[9] = 'bipolar'

    # 얻은 값들 엑셀 쓰기
    max_col = s_curve.max_column
    for row in range(0, len(curve_info)):
        s_curve.cell(row+1, max_col+1).value = curve_info[row]

    # file name change
    old_file = os.path.join(DIR, f'{img_file_name}')
    new_file = os.path.join(DIR, f'{curve_info[0]}.png')
    os.rename(old_file, new_file)

    wb.save(DIR + '/image.xlsx')
    wb.close()
    return print(f'{image_number}번째 이미지 done')

# element 개수마다 element_info 추가됨
def device(paper_name, image_number, DIR, img_file_name, figure_number):
    print('device 그래프')

    wb = openpyxl.load_workbook(DIR + '/image.xlsx')
    s_device = wb['s_device']

    info = ['file_name', 'figure_name', 'device']

    # element 개수만큼 생김
    # 0 : figure_name / 1 : device / [2: element / 3: thickness]
    device_info = ['NAN'] * 2
    device_info[0] = 'figure.' + input('Figure 이름 : ')

    element_cnt = int(input('element 총 몇 개? : '))

    fig_check = input('이전 figure와 다른 그래프 / 같은 figure인가요?(y/n) : ')
    if fig_check == 'y' : figure_number += 1
    else : figure_number = 1

    i_number = image_number
    image_name = f'{paper_name}.device{i_number}_{figure_number}'

    device_info[1] = input('Device 구성 : ')

    for i in range(element_cnt):
        element_info = ['NAN'] * 2
        element_info[0] = input(f'{i+1} 번째 소재 element : ')
        if input('두께 계산해야함? (y/n) : ') == 'y':
            element_info[1] = thickness()
        else: element_info[1] = input('두께 : ')
        device_info.append(f'[{element_info[0]}, {element_info[1]}]')
        info.append(f'layer_{i+1}')
    device_info.insert(0, image_name)

    # 얻은 값들 엑셀 쓰기
    max_col = s_device.max_column
    for row in range(0, len(device_info)):
        s_device.cell(row+1, max_col+1).value = device_info[row]

    # index 정보 쓰기
    for i in range(0, len(info)):
        s_device.cell(i+1, 1).value = info[i]

    # file name change
    old_file = os.path.join(DIR, f'{img_file_name}')
    new_file = os.path.join(DIR, f'{device_info[0]}.png')
    os.rename(old_file, new_file)

    wb.save(DIR + '/image.xlsx')
    wb.close()
    return print(f'{image_number}번째 이미지 done')

def endurance(paper_name, image_number, DIR, img_file_name, figure_number):
    print('Endurance 그래프')

    wb = openpyxl.load_workbook(DIR + '/image.xlsx')
    s_endurance = wb['s_endurance']

    i_number = image_number
    line_cnt = int(input('선 총 몇개? : '))

    # 0 : file name / 1 : figure_name / [2: x-axis / 3: y-axis]
    # 4 : Device / 5: Cycle / 6 : temperature
    # 7 : high resistance / 8 : high gradient
    # 9 : low resistance / 10 : low gradient

    info = ['file_name', 'figure_name', 'x-axis', 'y-axis',
            'Device', 'temperature', 'cycles',
            'high resistance', 'high gradient', 'low resistance', 'low gradient']

    for i in range(0, len(info)):
        s_endurance.cell(i+1, 1).value = info[i]

    enudrance_set_info = ['NAN'] * 6
    enudrance_set_info[0] = 'figure.' + input('Figure 이름 : ')

    x_first = input('x-axis first : ')
    x_last = input('x-axis last : ')
    enudrance_set_info[1] = f'[{x_first},{x_last}]'

    y_first = input('y-axis first : ')
    y_last = input('y-axis last : ')
    enudrance_set_info[2] = f'[{y_first},{y_last}]'

    enudrance_set_info[3] = input('device : ')

    enudrance_set_info[4] = float(input('Temperature(K) : '))
    temp_check = input('절대 온도인가요? (y / n) : ')
    if temp_check == 'n':
        enudrance_set_info[4] += 273
    elif temp_check == 'y': enudrance_set_info[4] = "3.0*10^2"

    enudrance_set_info[5] = input('Cycle : ').replace('^', '**')
    time = float(eval(enudrance_set_info[5]))

    # 선 개수만큼 반복
    for i in range(line_cnt):
        print(f'{i+1}번째 선')

        endurance_info = ['NAN'] * 4

        # high resistance
        high_res_high = eval(input('위 그래프에 제일 오른쪽 끝 y값 : ').replace('^', '**'))
        high_res_low = eval(input('위 그래프에서 제일 왼쪽 끝 쪽 y값 : ').replace('^', '**'))
        endurance_info[0] = round((high_res_high + high_res_low) / 2, 2)
        # endurance_info[0] = str(endurance_info[0]).replace('E', '*10^')

        # high restance gradient
        endurance_info[1] = round((high_res_high - high_res_low) / float(time),2)
        # endurance_info[1] = str(endurance_info[1]).replace('E', '*10^')

        # low resistance
        low_res_high = eval(input('아래 그래프에서 제일 오른쪽 끝 y값 : ').replace('^', '**'))
        low_res_low = eval(input('아래 그래프에서 제일 왼쪽 끝  y값 : ').replace('^', '**'))
        endurance_info[2] = round((low_res_high + low_res_low) / 2, 2)
        # endurance_info[2] = str(endurance_info[2]).replace('E', '*10^')

        # low resistance gradient
        endurance_info[3] = round((low_res_high - low_res_low) / float(time), 2)
        # endurance_info[3] = str(endurance_info[3]).replace('E', '*10^')

        # file_name
        enudrance_set_info.insert(0, f'{paper_name}.retention{i_number}_{i+1}')

        # 고정 데이터와 합치기
        endurance_info = enudrance_set_info + endurance_info

        # file name change
        old_file = os.path.join(DIR, f'{img_file_name}')
        new_file = os.path.join(DIR, f'{paper_name}.endurance{i_number}_{i+1}.png')
        os.rename(old_file, new_file)

        # 얻은 값들 엑셀 쓰기
        max_col = s_endurance.max_column

        for row in range(1, len(endurance_info)+1):
            s_endurance.cell(row, max_col + 1).value = endurance_info[row-1]

        wb.save(DIR + '/image.xlsx')
    wb.close()
    return print(f'{image_number}번째 이미지 done')

def retention(paper_name, image_number, DIR, img_file_name, figure_number):
    print('Retention 그래프')

    wb = openpyxl.load_workbook(DIR + '/image.xlsx')
    s_retention = wb['s_retention']

    i_number = image_number
    line_cnt = int(input('선 총 몇개? : '))

    info = ['file_name', 'figure_name', 'x-axis', 'y-axis',
            'Device', 'time', 'temperature',
            'high resistance', 'high gradient', 'low resistance', 'low gradient']

    for i in range(0,len(info)):
        s_retention.cell(i+1, 1).value = info[i]

    retention_set_info = ['NAN'] * 6
    retention_set_info[0] = 'figure.' + input('Figure 이름 : ')

    x_first = input('x-axis first : ')
    x_last = input('x-axis last : ')
    retention_set_info[1] = f'[{x_first},{x_last}]'

    y_first = input('y-axis first : ')
    y_last = input('y-axis last : ')
    retention_set_info[2] = f'[{y_first},{y_last}]'

    retention_set_info[3] = input('device : ')
    retention_set_info[4] = input('time : ').replace('^', '**')
    cycle = float(eval(retention_set_info[4]))

    retention_set_info[5] = float(input('Temperature(K) : '))
    temp_check = input('절대 온도인가요? (y / n) : ')
    if temp_check == 'n': retention_set_info[5] += 273
    elif temp_check == 'y': retention_set_info[5] = "3.0*10^2"


# 선 개수만큼 반복
    for i in range(line_cnt):
        print(f'{i+1}번째 선')

        retention_info = ['NAN'] * 4

        # high resistance
        high_res_high = eval(input('위 그래프에 제일 높은 y값 : ').replace('^', '**'))
        high_res_low = eval(input('위 그래프에서 제일 낮은 y값 : ').replace('^', '**'))
        retention_info[0] = round((high_res_high + high_res_low) / 2, 2)

        # high restance gradient
        retention_info[1] = round((high_res_high - high_res_low) / float(cycle), 2)
        # retention_info[1] = str(retention_info[1]).replace('E', '*10^')

        # low resistance
        low_res_high = eval(input('아래 그래프에서 제일 높은 y값 : ').replace('^', '**'))
        low_res_low = eval(input('아래 그래프에서 제일 낮은 y값 : ').replace('^', '**'))
        retention_info[2] = round((low_res_high + low_res_low) / 2, 2)
        # retention_info[2] = str(retention_info[2]).replace('E', '*10^')

        # low resistance gradient
        retention_info[3] = round((low_res_high - low_res_low) / float(cycle), 2)
        # retention_info[3] = str(retention_info[3]).replace('E', '*10^')

        # file_name
        retention_set_info.insert(0, f'{paper_name}.retention{i_number}_{i+1}')

        # file name change
        old_file = os.path.join(DIR, f'{img_file_name}')
        new_file = os.path.join(DIR, f'{paper_name}.retention{i_number}_{i+1}.png')
        os.rename(old_file, new_file)

        # 고정 데이터와 합치기
        retention_info = retention_set_info + retention_info

        # 얻은 값들 엑셀 쓰기
        max_col = s_retention.max_column

        for row in range(1, len(retention_info)+1):
            s_retention.cell(row, max_col + 1).value = retention_info[row-1]

        wb.save(DIR + '/image.xlsx')
    wb.close()
    return print(f'{image_number}번째 이미지 done')

def process(name, DIR):
    # 이미지 데이터 엑셀 생성
    wb = openpyxl.Workbook()

    # 엑셀 시트 총 4개
    s_curve = wb.active
    s_curve.title = 's_curve'
    wb.create_sheet('s_device')
    wb.create_sheet('s_endurance')
    wb.create_sheet('s_retention')

    wb.save(DIR + '/image.xlsx')

    paper_name = name
    image_cnt = [0] * 4
    figure_number = 0

    image_list = (os.listdir(DIR))

    try :
        if '.DS_Store' in image_list : image_list.remove('.DS_Store')
        if 'image.xlsx' in image_list : image_list.remove('image.xlsx')
    except : print('프로그램 시작')

    image_list = natsort.natsorted(image_list)
    print(image_list)

    for image in range(len(image_list) + 1):
        print('0 : I-V curve / 1 : device / 2 : endurance(cycle) / 3 : retention(time)')
        image_type = input("Image 타입 입력해주세요 : ")

        if image_type == '0':
            image_cnt[0] += 1
            curve(paper_name, str(image_cnt[0]), DIR, image_list[image], figure_number)
        elif image_type == '1':
            image_cnt[1] += 1
            device(paper_name, str(image_cnt[1]), DIR, image_list[image], figure_number)
        elif image_type == '2':
            image_cnt[2] += 1
            endurance(paper_name, str(image_cnt[2]), DIR, image_list[image], figure_number)
        elif image_type == '3':
            image_cnt[3] += 1
            retention(paper_name, str(image_cnt[3]), DIR, image_list[image], figure_number)

AB_DIR = '/Users/SBJ/Desktop/kriss/thesis'

# 총 3개의 논문, [0,1] 번째는 이미지, QA / 마지막 논문은 text라 해당 안됨
file_list = (os.listdir(AB_DIR))
if file_list[0] or file_list[-1] == '.DS_Store' or 'TextImg':
    file_list.remove('.DS_Store')
    file_list.remove('TextImg')

file_list.sort(key=float)

print(file_list)
DIR = f'{AB_DIR}/{file_list[0]}'

for i in range(2):
    print(file_list[i], DIR)
    process(file_list[i], DIR)