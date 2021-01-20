# 2021/01/01
# Tag 된 단어들 Numeric data로 만들기
# 엑셀은 WordTag에서 만들어진 엑셀을 사용해 미리 포맷되어 있어야함(TextJoin)
# 짝 묶어주고, Numeric 짝 없는건 x 채우는건 수동

import openpyxl
import re

from os import listdir
from os.path import isfile, join

# category_info
material = ['m', 'mp', 'md', 'mdp']
device = ['d', 'ds', 'da',
          'dd', 'dds', 'dda',
          'md', 'mds', 'mda']
performance = ['pp', 'pc', 'pv', 'pr', 'po', 'ps', 'pe', 'pab', 'plec',
               'dpp', 'dpc', 'dpv', 'dpr', 'dpo', 'dps', 'dpe', 'dpab', 'dplec',
               'mpp', 'mpc', 'mpv', 'mpr', 'mpo', 'mps', 'mpe', 'mpab', 'mplec']
mechanism = ['mc', ' mmc', 'dmc']
environment = ['e', 'me', 'de']
synthesis = ['s']

number = ['n', 'mn']
metric = ['u', 'mu']

category_rule = {'0': material, '1': device, '2': performance, '3': mechanism, '4': environment, '5': synthesis}
# number_trim
trim = re.compile("[^0-9]")

# 절대 경로 / 상대 경로
AB_DIR = '/Users/SBJ/Desktop/kriss/Data/Text/Word'
DIR = 'Data/Text/Word/'

# Data/Text/Word 에 있는 파일을 리스트로 생성
file_list = [f for f in listdir(AB_DIR) if isfile(join(AB_DIR, f))]
print(file_list)

wb = openpyxl.load_workbook(DIR + file_list[1])
sheet = wb.active

max_row = sheet.max_row
max_col = sheet.max_column

for row in range(3, max_row + 1, 2):
    numerical, numerical_temp = [], []
    category = [0] * 6

    for col in range(4, max_col + 1):
        # 숫자 다 뺌
        cell_info = sheet.cell(row, col).value
        if not type(cell_info) == str : break
        cell_info = ("".join(trim.findall(sheet.cell(row, col).value)))

        if cell_info in number or cell_info in metric:
            numerical_temp.append(sheet.cell(row-1, col).value)
            if not len(numerical_temp) % 2 : numerical.append(numerical_temp)
        for i in range(0, len(category_rule)):
            if cell_info in category_rule[str(i)] : category[i] += 1
    if len(numerical_temp) % 2 : numerical.append(numerical_temp + ['x'])
    sheet.cell(row, 2).value = str(category)
    sheet.cell(row, 3).value = str(numerical)

wb.save(DIR + file_list[1])