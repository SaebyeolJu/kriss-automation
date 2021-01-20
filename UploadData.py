# -*- coding: utf-8 -*-
# upload.py
# automate filling in web forms(Kriss)
import time
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait

def login():
    url = "https://www.station.re.kr/group/kriss-materials-data-center/data-management"
    driver.get(url)

    # login process
    driver.find_element_by_id('_com_liferay_login_web_portlet_LoginPortlet_login').send_keys('ki0086')
    driver.find_element_by_id('_com_liferay_login_web_portlet_LoginPortlet_password').send_keys('evoid8913')
    driver.find_element_by_css_selector('button.gradient-btn.col-md-12.h-100.border-0').click()

def downToTree():
    # after login, switch to iframe
    # there are two iframes(1. file_tree / 2. data info edit )

    time.sleep(5)
    iframes = driver.find_elements_by_tag_name('iframe')

    # processing file_tree iframe
    tree_iframe = iframes[0]
    driver.switch_to.frame(tree_iframe)

    # expanding the tree
    tree = 'fancytree-node fancytree-folder fancytree-lastsib fancytree-exp-nl fancytree-ico-cf'
    tree_element = driver.find_element_by_xpath('//*[@id="ft-id-1"]/li/span')
    webdriver.ActionChains(driver).move_to_element(tree_element).click(tree_element).perform()

def files():
    AB_DIR = '/Users/SBJ/Desktop/kriss/thesis'
    folder_list = (os.listdir(AB_DIR))

    if folder_list[0] or folder_list[-1] == '.DS_Store' or 'TextImg':
        folder_list.remove('.DS_Store')
        folder_list.remove('TextImg')

    folder_list = sorted(folder_list)
    DIR = AB_DIR + '/' + str(folder_list[0])

    image_list = [each for each in os.listdir(DIR) if each.endswith('.png')]
    device_img, iv_img, retention_img, endurance_img = [], [], [], []
    for image in image_list:
        if 'device' in image: device_img.append(image)
        elif 'IV' in image: iv_img.append(image)
        elif 'retention' in image: retention_img.append(image)
        elif 'endurance' in image: endurance_img.append(image)

    QA_file = [each for each in os.listdir(DIR) if 'Q_A' in each]
    para_file = [each for each in os.listdir(DIR) if 'paragraph_labeling' in each]
    text_label_file = [each for each in os.listdir(DIR) if 'text' in each]

    file_dic = {'QA': QA_file[0], 'device': sorted(device_img), 'iv': sorted(iv_img),
                'retention': sorted(retention_img), 'endurance': sorted(endurance_img),
                'para': para_file, 'text': text_label_file, 'DIR': DIR}

    return file_dic

def QA(DIR, QA_file):
    print('QA 시작')
    downToTree()

    time.sleep(1)
    QA_tree_1 = driver.find_element_by_xpath('//*[@id="ui-id-1"]/ul/li[2]')
    webdriver.ActionChains(driver).move_to_element(QA_tree_1).click(QA_tree_1).perform()

    time.sleep(1)
    QA_tree_2 = driver.find_element_by_xpath('//*[@id="ui-id-2"]/ul/li/span')
    webdriver.ActionChains(driver).move_to_element(QA_tree_2).click(QA_tree_2).perform()
    # QA 경로 가기 끝남

    driver.switch_to.default_content()

    info_frame = driver.find_element_by_id('_OSPVisualizing_analyzer_DataInfo_INSTANCE_LAYOUT_canvas')
    driver.switch_to.frame(info_frame)

    # hit the button
    # button tag 2개고, id값 새로고침 할 때마다 바뀜 (Xpath, id값 절대 쓰면 X)
    time.sleep(1)
    buttons = driver.find_elements_by_tag_name('button')

    time.sleep(3)
    ActionChains(driver).move_to_element(buttons[2]).click(buttons[2]).perform()

    # 이미지 올리기
    driver.find_element_by_id("_OSPVisualizing_analyzer_DataInfo_dataFile").send_keys(f'{DIR}/{QA_file}')

    # frame 변경
    input_frame = driver.find_element_by_css_selector('#_OSPVisualizing_analyzer_DataInfo_sdeMetaDataIframe_1')
    driver.switch_to.frame(input_frame)

    # input
    wb = openpyxl.load_workbook(DIR + '/info.xlsx')
    sheet = wb['Sheet']
    max_row = sheet.max_row

    # 데이터 입력 폼
    input_space = driver.find_elements_by_css_selector("input[type='text']")

    for i in range(1, max_row):
        cell = 'A' + str(i)
        data = sheet[cell].value

        input_space[i-1].clear()
        input_space[i-1].send_keys(data)

    driver.switch_to.default_content()
    driver.switch_to.frame(info_frame)

    save_btn = driver.find_elements_by_tag_name('button')
    ActionChains(driver).move_to_element(save_btn[3]).click(save_btn[3]).perform()
    time.sleep(3)
    print("QA_summit done")

def para(DIR, para_file):
    print("paragraph start")
    downToTree()

    time.sleep(1)
    text_tree = driver.find_element_by_xpath('//*[@id="ui-id-1"]/ul/li[3]/span/span[2]')
    webdriver.ActionChains(driver).move_to_element(text_tree).click(text_tree).perform()

    time.sleep(1)
    para_tree = driver.find_element_by_xpath('//*[@id="ui-id-2"]/ul/li[1]/span/span[2]')
    webdriver.ActionChains(driver).move_to_element(para_tree).click(para_tree).perform()

    driver.switch_to.default_content()

    info_frame = driver.find_element_by_id('_OSPVisualizing_analyzer_DataInfo_INSTANCE_LAYOUT_canvas')
    driver.switch_to.frame(info_frame)

    # hit the button
    # button tag 2개고, id값 새로고침 할 때마다 바뀜 (Xpath, id값 절대 쓰면 X)
    time.sleep(1)
    buttons = driver.find_elements_by_tag_name('button')

    time.sleep(1)
    ActionChains(driver).move_to_element(buttons[2]).click(buttons[2]).perform()

    # 이미지 올리기
    driver.find_element_by_id("_OSPVisualizing_analyzer_DataInfo_dataFile").send_keys(f'{DIR}/{para_file[0]}')

    # frame 변경
    input_frame = driver.find_element_by_css_selector('#_OSPVisualizing_analyzer_DataInfo_sdeMetaDataIframe_1')
    driver.switch_to.frame(input_frame)

    # input
    wb = openpyxl.load_workbook(DIR + '/info.xlsx')
    sheet = wb['Sheet']
    max_row = sheet.max_row

    # 데이터 입력 폼
    input_space = driver.find_elements_by_css_selector("input[type='text']")

    for i in range(1, max_row):
        cell = 'A' + str(i)
        data = sheet[cell].value

        input_space[i-1].clear()
        input_space[i-1].send_keys(data)

    driver.switch_to.default_content()
    driver.switch_to.frame(info_frame)

    save_btn = driver.find_elements_by_tag_name('button')
    ActionChains(driver).move_to_element(save_btn[3]).click(save_btn[3]).perform()
    time.sleep(3)
    print("paragraph done")

def text(DIR, text_file):
    print("text start")

    downToTree()

    time.sleep(1)
    text_tree = driver.find_element_by_xpath('//*[@id="ui-id-1"]/ul/li[3]/span')
    webdriver.ActionChains(driver).move_to_element(text_tree).click(text_tree).perform()

    time.sleep(1)
    text_label_tree = driver.find_element_by_xpath('//*[@id="ui-id-2"]/ul/li[2]/span/span[2]')
    webdriver.ActionChains(driver).move_to_element(text_label_tree).click(text_label_tree).perform()

    driver.switch_to.default_content()

    info_frame = driver.find_element_by_id('_OSPVisualizing_analyzer_DataInfo_INSTANCE_LAYOUT_canvas')
    driver.switch_to.frame(info_frame)

    # hit the entry button
    time.sleep(1)
    buttons = driver.find_elements_by_tag_name('button')

    time.sleep(1)
    ActionChains(driver).move_to_element(buttons[2]).click(buttons[2]).perform()

    # 이미지 올리기
    driver.find_element_by_id("_OSPVisualizing_analyzer_DataInfo_dataFile").send_keys(f'{DIR}/{text_file[0]}')

    # frame 변경
    input_frame = driver.find_element_by_css_selector('#_OSPVisualizing_analyzer_DataInfo_sdeMetaDataIframe_1')
    driver.switch_to.frame(input_frame)

    # input
    wb = openpyxl.load_workbook(DIR + '/info.xlsx')
    sheet = wb['Sheet']
    max_row = sheet.max_row

    # 데이터 입력 폼
    input_space = driver.find_elements_by_css_selector("input[type='text']")

    for i in range(1, max_row+1):
        cell = 'A' + str(i)
        data = sheet[cell].value

        input_space[i-1].clear()
        input_space[i-1].send_keys(data)

    driver.switch_to.default_content()
    driver.switch_to.frame(info_frame)

    save_btn = driver.find_elements_by_tag_name('button')
    ActionChains(driver).move_to_element(save_btn[3]).click(save_btn[3]).perform()
    time.sleep(3)
    print("text done")

def iv(DIR, iv_img):
    print("curve start")
    downToTree()

    # read exel
    wb = openpyxl.load_workbook(DIR + '/image.xlsx')
    s_curve = wb['s_curve']
    max_row = s_curve.max_row

    time.sleep(1)
    image_element = driver.find_element_by_xpath('//*[@id="ui-id-1"]/ul/li[1]/span/span[2]')
    webdriver.ActionChains(driver).move_to_element(image_element).click(image_element).perform()

    time.sleep(1)
    iv_tree = driver.find_element_by_xpath('//*[@id="ui-id-2"]/ul/li[3]/span')
    webdriver.ActionChains(driver).move_to_element(iv_tree).click(iv_tree).perform()

    for i in range(len(iv_img)):
        if i > 0:
            time.sleep(3)
            driver.switch_to.default_content()
            iframes = driver.find_elements_by_tag_name('iframe')

            # processing file_tree iframe
            tree_iframe = iframes[0]

            driver.switch_to.frame(tree_iframe)
            webdriver.ActionChains(driver).move_to_element(iv_tree).click(iv_tree).perform()

        time.sleep(1)
        driver.switch_to.default_content()

        time.sleep(1)
        info_frame = driver.find_element_by_id('_OSPVisualizing_analyzer_DataInfo_INSTANCE_LAYOUT_canvas')
        driver.switch_to.frame(info_frame)

        # hit the entry button
        time.sleep(1)
        buttons = driver.find_elements_by_tag_name('button')
        ActionChains(driver).move_to_element(buttons[2]).click(buttons[2]).perform()

        time.sleep(1)
        print(f'{DIR}/{iv_img[i]}')
        driver.find_element_by_id("_OSPVisualizing_analyzer_DataInfo_dataFile").send_keys(f'{DIR}/{iv_img[i]}')

        time.sleep(1)
        input_frame = driver.find_element_by_css_selector('#_OSPVisualizing_analyzer_DataInfo_sdeMetaDataIframe_1')
        driver.switch_to.frame(input_frame)

        # 데이터 입력 폼
        input_space = driver.find_elements_by_css_selector("input[type='text']")

        for j in range(0, max_row-1):
            cell = chr(i+66) + str(j+2)
            data = s_curve[cell].value
            if type(data) != int:
                data = data.replace("=", "")
            input_space[j].clear()
            input_space[j].send_keys(data)

        driver.switch_to.default_content()
        driver.switch_to.frame(info_frame)

        save_btn = driver.find_elements_by_tag_name('button')
        ActionChains(driver).move_to_element(save_btn[3]).click(save_btn[3]).perform()
        print(f'{i+1}번째 iv 이미지 done')
        time.sleep(3)
    time.sleep(3)
    print("iv curve done")

def endurance(DIR, endurance_img):
    print("endurance 시작")
    downToTree()

    # read exel
    wb = openpyxl.load_workbook(DIR + '/image.xlsx')
    s_endurance = wb['s_endurance']
    max_row = s_endurance.max_row

    time.sleep(1)
    image_element = driver.find_element_by_xpath('//*[@id="ui-id-1"]/ul/li[1]/span/span[2]')
    webdriver.ActionChains(driver).move_to_element(image_element).click(image_element).perform()

    time.sleep(2)
    endurance_tree = driver.find_element_by_xpath('//*[@id="ui-id-2"]/ul/li[2]/span')
    webdriver.ActionChains(driver).move_to_element(endurance_tree).click(endurance_tree).perform()

    for i in range(len(endurance_img)):
        if i > 0:
            time.sleep(3)
            driver.switch_to.default_content()
            iframes = driver.find_elements_by_tag_name('iframe')

            # processing file_tree iframe
            tree_iframe = iframes[0]

            driver.switch_to.frame(tree_iframe)
            webdriver.ActionChains(driver).move_to_element(endurance_tree).click(endurance_tree).perform()

        time.sleep(1)
        driver.switch_to.default_content()

        time.sleep(1)
        info_frame = driver.find_element_by_id('_OSPVisualizing_analyzer_DataInfo_INSTANCE_LAYOUT_canvas')
        driver.switch_to.frame(info_frame)

        time.sleep(1)
        buttons = driver.find_elements_by_tag_name('button')
        ActionChains(driver).move_to_element(buttons[2]).click(buttons[2]).perform()

        # 이미지 올리기
        print(f'{DIR}/{endurance_img[i]}')
        driver.find_element_by_id("_OSPVisualizing_analyzer_DataInfo_dataFile").send_keys(f'{DIR}/{endurance_img[i]}')

        # frame 변경
        input_frame = driver.find_element_by_css_selector('#_OSPVisualizing_analyzer_DataInfo_sdeMetaDataIframe_1')
        driver.switch_to.frame(input_frame)

        # 데이터 입력 폼
        input_space = driver.find_elements_by_css_selector("input[type='text']")

        for j in range(0, max_row-1):
            cell = chr(i+66) + str(j+2)
            data = s_endurance[cell].value
            if type(data) != int:
                data = data.replace("=","")

            input_space[j].clear()
            input_space[j].send_keys(data)

        driver.switch_to.default_content()
        driver.switch_to.frame(info_frame)

        save_btn = driver.find_elements_by_tag_name('button')
        ActionChains(driver).move_to_element(save_btn[3]).click(save_btn[3]).perform()
        print(f'{i+1}번째 endurance 이미지 done')
        time.sleep(3)
    time.sleep(3)
    print("endurance done")

def retention(DIR, retention_img):
    print('retention 이미지 시작')
    downToTree()

    # xlsx read
    wb = openpyxl.load_workbook(DIR + '/image.xlsx')
    s_retention = wb['s_retention']
    max_row = s_retention.max_row

    time.sleep(2)
    image_element = driver.find_element_by_xpath('//*[@id="ui-id-1"]/ul/li[1]/span/span[2]')
    webdriver.ActionChains(driver).move_to_element(image_element).click(image_element).perform()

    time.sleep(1)
    retention_tree = driver.find_element_by_xpath('//*[@id="ui-id-2"]/ul/li[4]/span')
    webdriver.ActionChains(driver).move_to_element(retention_tree).click(retention_tree).perform()

    driver.switch_to.default_content()

    info_frame = driver.find_element_by_id('_OSPVisualizing_analyzer_DataInfo_INSTANCE_LAYOUT_canvas')
    driver.switch_to.frame(info_frame)

    for i in range(len(retention_img)):
        if i > 0:
            time.sleep(3)
            driver.switch_to.default_content()
            iframes = driver.find_elements_by_tag_name('iframe')

            # processing file_tree iframe
            tree_iframe = iframes[0]

            driver.switch_to.frame(tree_iframe)
            webdriver.ActionChains(driver).move_to_element(retention_tree).click(retention_tree).perform()

        time.sleep(1)
        driver.switch_to.default_content()

        time.sleep(1)
        info_frame = driver.find_element_by_id('_OSPVisualizing_analyzer_DataInfo_INSTANCE_LAYOUT_canvas')
        driver.switch_to.frame(info_frame)

        # hit the button
        time.sleep(1)
        buttons = driver.find_elements_by_tag_name('button')
        ActionChains(driver).move_to_element(buttons[2]).click(buttons[2]).perform()

        # 이미지 올리기
        time.sleep(1)
        print(f'{DIR}/{retention_img[i]}')
        driver.find_element_by_id("_OSPVisualizing_analyzer_DataInfo_dataFile").send_keys(f'{DIR}/{retention_img[i]}')

        # frame 변경
        time.sleep(1)
        input_frame = driver.find_element_by_css_selector('#_OSPVisualizing_analyzer_DataInfo_sdeMetaDataIframe_1')
        driver.switch_to.frame(input_frame)

        # 데이터 입력 폼
        input_space = driver.find_elements_by_css_selector("input[type='text']")

        for j in range(0, max_row-1):
            cell = chr(i+66) + str(j+2)
            data = s_retention[cell].value
            if type(data) != int:
                data = data.replace("=","")

            input_space[j].clear()
            input_space[j].send_keys(data)

        driver.switch_to.default_content()
        driver.switch_to.frame(info_frame)

        save_btn = driver.find_elements_by_tag_name('button')
        ActionChains(driver).move_to_element(save_btn[3]).click(save_btn[3]).perform()
        print(f'{i+1}번째 retention 완료')
        time.sleep(3)
    time.sleep(3)
    print("retention 완료")

def device(DIR, device_img):
    print('device 이미지 시작')
    downToTree()

    # xlsx read
    wb = openpyxl.load_workbook(DIR + '/image.xlsx')
    s_device = wb['s_device']
    max_row = s_device.max_row

    image_element = driver.find_element_by_xpath('//*[@id="ui-id-1"]/ul/li[1]/span/span[2]')
    webdriver.ActionChains(driver).move_to_element(image_element).click(image_element).perform()

    time.sleep(1)
    device_tree = driver.find_element_by_xpath('//*[@id="ui-id-2"]/ul/li[1]/span/span[2]')
    webdriver.ActionChains(driver).move_to_element(device_tree).click(device_tree).perform()

    driver.switch_to.default_content()

    info_frame = driver.find_element_by_id('_OSPVisualizing_analyzer_DataInfo_INSTANCE_LAYOUT_canvas')
    driver.switch_to.frame(info_frame)

    for i in range(len(device_img)):
        if i > 0:
            time.sleep(3)
            driver.switch_to.default_content()
            iframes = driver.find_elements_by_tag_name('iframe')

            # processing file_tree iframe
            tree_iframe = iframes[0]

            driver.switch_to.frame(tree_iframe)
            webdriver.ActionChains(driver).move_to_element(device_tree).click(device_tree).perform()

        time.sleep(1)
        driver.switch_to.default_content()

        time.sleep(1)
        info_frame = driver.find_element_by_id('_OSPVisualizing_analyzer_DataInfo_INSTANCE_LAYOUT_canvas')
        driver.switch_to.frame(info_frame)

        # hit the entry button
        time.sleep(1)
        buttons = driver.find_elements_by_tag_name('button')
        ActionChains(driver).move_to_element(buttons[2]).click(buttons[2]).perform()

        # 이미지 올리기
        print(f'{DIR}/{device_img[i]}')
        driver.find_element_by_id("_OSPVisualizing_analyzer_DataInfo_dataFile").send_keys(f'{DIR}/{device_img[i]}')

        # text-field frame 변경
        input_frame = driver.find_element_by_css_selector('#_OSPVisualizing_analyzer_DataInfo_sdeMetaDataIframe_1')
        driver.switch_to.frame(input_frame)

        # 데이터 입력 폼
        input_space = driver.find_elements_by_css_selector("input[type='text']")

        for j in range(0, max_row-1):
            cell = chr(i+66) + str(j+2)
            data = s_device[cell].value
            if type(data) != int:
                data = data.replace("=", "")

            input_space[j].clear()
            input_space[j].send_keys(data)

        driver.switch_to.default_content()
        driver.switch_to.frame(info_frame)

        save_btn = driver.find_elements_by_tag_name('button')
        ActionChains(driver).move_to_element(save_btn[3]).click(save_btn[3]).perform()
        print(f'{i+1}번째 device 완료')
    time.sleep(3)
    print("device 완료")

options = webdriver.ChromeOptions()
driver = webdriver.Chrome('Data/chromedriver')

login()

# step :
# 1. down to under the tree -> 2.Entry add -> 3. fill requirements : meta data, file
file_dic = files()

methods = {'QA': QA, 'para': para, 'text': text,
           'device': device, 'retention': retention, 'iv': iv, 'endurance': endurance}

print(file_dic)

for name, file in file_dic.items():
    if name == 'DIR': break
    elif len(file) > 0:
        print(name, file, methods[name])
        driver.refresh()
        methods[name](file_dic['DIR'], file)

print("<< 모든 input 완료 >>")
