# -*- coding: utf-8 -*-

import os
from os import listdir
from os.path import isfile, join

import openpyxl

AB_DIR = '/Users/SBJ/Desktop/kriss/Data/Image/Device'
DIR = 'Data/Image/Device/'

# '.DS_Store' 주의
# file 개수만큼 반복
file_list = [f for f in listdir(AB_DIR) if isfile(join(AB_DIR, f))]
print(file_list)

wb = openpyxl.load_workbook(DIR + file_list[1])
sheet = wb.active

# column 개수 - 1 만큼 업로드 해야함
# 채워야 하는 데이터 input : 각 column의 row 개수
max_row = sheet.max_row
max_col = sheet.max_column

data_list = {}
data = []

for col in range(2, max_col + 1):
    for row in range(1, max_row +1):
        if sheet.cell(row = row, column = col).value :
            data.append(sheet.cell(row = row, column = col).value)
    data_list[col-1] = data
    data = []
print(data_list)

wb.close()